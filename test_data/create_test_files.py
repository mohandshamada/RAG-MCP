"""Create test files for testing the RAG server"""
from PyPDF2 import PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from pathlib import Path

# Create test_data directory
test_dir = Path("test_data")
test_dir.mkdir(exist_ok=True)

# Create a simple PDF with text using reportlab
try:
    from reportlab.pdfgen import canvas
    pdf_path = test_dir / "test_document.pdf"
    c = canvas.Canvas(str(pdf_path), pagesize=letter)
    c.drawString(100, 750, "Test Document for RAG Server")
    c.drawString(100, 730, "")
    c.drawString(100, 710, "This is a test document containing sample text.")
    c.drawString(100, 690, "It includes information about artificial intelligence.")
    c.drawString(100, 670, "Machine learning is a subset of AI.")
    c.drawString(100, 650, "Natural language processing helps computers understand text.")
    c.drawString(100, 630, "This document is used for testing PDF ingestion.")
    c.save()
    print(f"[OK] Created PDF: {pdf_path}")
except ImportError:
    print("! reportlab not installed, creating blank PDF")
    pdf_path = test_dir / "test_document.pdf"
    pdf = PdfWriter()
    pdf.add_blank_page(width=612, height=792)
    with open(pdf_path, 'wb') as f:
        pdf.write(f)
    print(f"[OK] Created blank PDF: {pdf_path}")

# Create a simple Excel file
excel_path = test_dir / "test_spreadsheet.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Data"
ws['A1'] = "Item"
ws['B1'] = "Value"
ws['A2'] = "Compliance Score"
ws['B2'] = 95
ws['A3'] = "Total Items"
ws['B3'] = 10
ws['A4'] = "Status"
ws['B4'] = "COMPLIANT"
wb.save(excel_path)
print(f"[OK] Created Excel: {excel_path}")

# Create a test requirements file
req_path = test_dir / "test_requirements.txt"
with open(req_path, 'w') as f:
    f.write("""Test Requirements Specification

1. The system must support PDF document processing
2. The system must provide semantic search capabilities
3. The system must generate compliance reports
4. Response time must be under 200ms for queries
5. The system must handle documents up to 500MB
""")
print(f"[OK] Created requirements: {req_path}")

print("\nAll test files created successfully!")
